import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter


class Vacancy:
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_currency = vacancy['salary_currency']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.year = int(vacancy['published_at'][:4])
        self.area_name = vacancy['area_name']


class DataSet:
    def __init__(self, file, vacancy):
        self.file_name = file
        self.vacancy_name = vacancy

    @staticmethod
    def get_average_dict(data):
        result = {}
        for key, data in data.items():
            result[key] = int(sum(data) / len(data))
        return result

    @staticmethod
    def increment(subject, key, value):
        if key in subject:
            subject[key] += value
        else:
            subject[key] = value

    def csv_reader(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            titles = next(reader)
            titles_count = len(titles)
            for row in reader:
                if '' not in row and len(row) == titles_count:
                    yield dict(zip(titles, row))

    def get_statistics(self):
        salary_of_vacancy_name = {}
        salary_city = {}
        vacancies_count = 0
        salary = {}
        for vacancy in self.csv_reader():
            vacancy = Vacancy(vacancy)
            self.increment(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(salary_of_vacancy_name, vacancy.year, [vacancy.salary_average])
            self.increment(salary_city, vacancy.area_name, [vacancy.salary_average])
            vacancies_count += 1
        vacancies_number = dict([(key, len(value)) for key, value in salary.items()])
        vacs_per_name = dict([(key, len(value)) for key, value in salary_of_vacancy_name.items()])
        if not salary_of_vacancy_name:
            salary_of_vacancy_name = dict([(key, [0]) for key, value in salary.items()])
            vacs_per_name = dict([(key, 0) for key, value in vacancies_number.items()])
        stat_salary = self.get_average_dict(salary)
        stat_salary_by_vac = self.get_average_dict(salary_of_vacancy_name)
        stat_salary_by_city = self.get_average_dict(salary_city)
        stat_salary_by_year = {}
        for year, salaries in salary_city.items():
            stat_salary_by_year[year] = round(len(salaries) / vacancies_count, 4)
        stat_salary_by_year = list(filter(lambda elem: elem[-1] >= 0.01,
                                          [(key, value) for key, value in stat_salary_by_year.items()]))
        stat_salary_by_year.sort(key=lambda elem: elem[-1], reverse=True)
        top_salary_by_year = stat_salary_by_year.copy()
        stat_salary_by_year = dict(stat_salary_by_year)
        stat_salary_by_city = list(filter(lambda elem: elem[0] in list(stat_salary_by_year.keys()),
                                          [(key, value) for key, value in stat_salary_by_city.items()]))
        stat_salary_by_city.sort(key=lambda elem: elem[-1], reverse=True)
        stat_salary_by_city = dict(stat_salary_by_city[:10])
        top_salary_by_year = dict(top_salary_by_year[:10])
        return stat_salary, vacancies_number, stat_salary_by_vac, vacs_per_name, stat_salary_by_city, top_salary_by_year

    @staticmethod
    def print_statistic(salary_by_year, vacs_per_year, salary_by_vac, count_by_vac, salary_by_city, city_percents):
        print('Динамика уровня зарплат по годам: {0}'.format(salary_by_year))
        print('Динамика количества вакансий по годам: {0}'.format(vacs_per_year))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(salary_by_vac))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(count_by_vac))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(salary_by_city))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(city_percents))


class Report:
    def __init__(self, vacancy, salary_by_year, vacs_per_year, salary_by_vac,
                 count_by_vac, salary_by_city, city_percents):
        self.wb = Workbook()
        self.vacancy_name = vacancy
        self.salary_by_year = salary_by_year
        self.vacs_per_year = vacs_per_year
        self.salary_by_vac = salary_by_vac
        self.count_by_vac = count_by_vac
        self.salary_by_city = salary_by_city
        self.city_percents = city_percents

    def create_xlsx_file(self):
        year_sheet = self.wb.active
        year_sheet.title = 'Статистика по годам'
        year_sheet.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name,
                           'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.salary_by_year.keys():
            year_sheet.append([year, self.salary_by_year[year], self.salary_by_vac[year],
                               self.vacs_per_year[year], self.count_by_vac[year]])
        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name,
                 ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]
        for i, width in enumerate(column_widths, 1):
            year_sheet.column_dimensions[get_column_letter(i)].width = width + 2
        data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city1, value1), (city2, value2) in zip(self.salary_by_city.items(), self.city_percents.items()):
            data.append([city1, value1, '', city2, value2])
        city_sheet = self.wb.create_sheet('Статистика по городам')
        for row in data:
            city_sheet.append(row)
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]
        for i, width in enumerate(column_widths, 1):
            city_sheet.column_dimensions[get_column_letter(i)].width = width + 2
        for col in 'ABCDE':
            year_sheet[col + '1'].font = Font(bold=True)
            city_sheet[col + '1'].font = Font(bold=True)
        for index, _ in enumerate(self.salary_by_city):
            city_sheet['E' + str(index + 2)].number_format = '0.00%'
        my_side = Side(border_style='thin', color='00000000')
        for row in range(len(data)):
            for col in 'ABDE':
                city_sheet[col + str(row + 1)].border = Border(right=my_side, top=my_side, left=my_side, bottom=my_side)
        self.salary_by_year[1] = 1
        for row, _ in enumerate(self.salary_by_year):
            for col in 'ABCDE':
                year_sheet[col + str(row + 1)].border = Border(left=my_side, bottom=my_side, right=my_side, top=my_side)
        self.wb.save('report.xlsx')


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')
        dataset = DataSet(self.file_name, self.vacancy_name)
        salary_by_year, vacs_per_year, salary_by_vac, count_by_vac, salary_by_city, city_percents\
            = dataset.get_statistics()
        dataset.print_statistic(salary_by_year, vacs_per_year, salary_by_vac,
                                count_by_vac, salary_by_city, city_percents)
        report = Report(self.vacancy_name, salary_by_year, vacs_per_year, salary_by_vac,
                        count_by_vac, salary_by_city, city_percents)
        report.create_xlsx_file()


if __name__ == '__main__':
    InputConnect()
    